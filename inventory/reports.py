from io import BytesIO

from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)

from openpyxl import Workbook
from openpyxl.styles import Font

from .models import Receipt

def receipt_pdf(request, pk):

    recepcion = get_object_or_404(
        Receipt,
        pk=pk
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = f'inline; filename="Recepcion_{recepcion.id}.pdf"'

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=(21*cm, 29.7*cm)
    )

    styles = getSampleStyleSheet()

    elementos = []

    elementos.append(
        Paragraph(
            "<b>MATERIAL CONTROL CM</b>",
            styles["Title"]
        )
    )

    elementos.append(
        Paragraph(
            f"<b>Recepción No.</b> {recepcion.id}",
            styles["Heading2"]
        )
    )

    elementos.append(Spacer(1, 0.5*cm))

    elementos.append(
        Paragraph(
            f"<b>Proyecto:</b> {recepcion.proyecto}",
            styles["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"<b>Proveedor:</b> {recepcion.proveedor}",
            styles["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"<b>Fecha:</b> {recepcion.fecha}",
            styles["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"<b>Remisión:</b> {recepcion.numero_remision}",
            styles["Normal"]
        )
    )

    elementos.append(Spacer(1, 0.5*cm))

    datos = [
        [
            "Código",
            "Material",
            "Unidad",
            "Cantidad"
        ]
    ]

    for detalle in recepcion.detalles.all():

        datos.append([
            detalle.material.codigo,
            detalle.material.nombre,
            detalle.material.unidad,
            detalle.cantidad,
        ])

    tabla = Table(datos)

    tabla.setStyle(

        TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.darkblue),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("GRID", (0,0), (-1,-1), 0.5, colors.black),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige),

            ("ALIGN", (0,0), (-1,-1), "CENTER"),

            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

        ])

    )

    elementos.append(tabla)

    doc.build(elementos)

    pdf = buffer.getvalue()

    buffer.close()

    response.write(pdf)

    return response

def receipt_excel(request, pk):

    recepcion = get_object_or_404(
        Receipt,
        pk=pk
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Recepción"

    titulo = "MATERIAL CONTROL CM - RECEPCIÓN"

    ws["A1"] = titulo
    ws["A1"].font = Font(
        bold=True,
        size=14
    )

    ws["A3"] = "Recepción No."
    ws["B3"] = recepcion.id

    ws["A4"] = "Proyecto"
    ws["B4"] = str(recepcion.proyecto)

    ws["A5"] = "Proveedor"
    ws["B5"] = str(recepcion.proveedor)

    ws["A6"] = "Fecha"
    ws["B6"] = str(recepcion.fecha)

    ws["A7"] = "Remisión"
    ws["B7"] = recepcion.numero_remision

    fila = 10

    encabezados = [

        "Código",

        "Material",

        "Unidad",

        "Cantidad",

        "Observación",

    ]

    for columna, encabezado in enumerate(encabezados, start=1):

        celda = ws.cell(
            row=fila,
            column=columna
        )

        celda.value = encabezado

        celda.font = Font(
            bold=True
        )

    fila += 1

    for detalle in recepcion.detalles.all():

        ws.cell(
            row=fila,
            column=1
        ).value = detalle.material.codigo

        ws.cell(
            row=fila,
            column=2
        ).value = detalle.material.nombre

        ws.cell(
            row=fila,
            column=3
        ).value = detalle.material.unidad

        ws.cell(
            row=fila,
            column=4
        ).value = float(detalle.cantidad)

        ws.cell(
            row=fila,
            column=5
        ).value = detalle.observacion

        fila += 1

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (

        f'attachment; filename="Recepcion_{recepcion.id}.xlsx"'

    )

    wb.save(response)

    return response    
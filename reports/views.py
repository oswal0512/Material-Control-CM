from django.shortcuts import render
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font

from materials.models import Material


def report_home(request):

    return render(
        request,
        "reports/home.html"
    )


def inventory_report(request):

    materiales = Material.objects.filter(
        activo=True
    ).order_by(
        "codigo"
    )

    return render(
        request,
        "reports/inventory.html",
        {
            "materiales": materiales
        }
    )


def inventory_excel(request):

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventario"

    encabezados = [

        "Código",
        "Material",
        "Unidad",
        "Stock",
        "Estado",

    ]

    for columna, titulo in enumerate(encabezados, start=1):

        celda = ws.cell(
            row=1,
            column=columna
        )

        celda.value = titulo

        celda.font = Font(
            bold=True
        )

    fila = 2

    materiales = Material.objects.filter(
        activo=True
    ).order_by(
        "codigo"
    )

    for material in materiales:

        if material.stock == 0:

            estado = "AGOTADO"

        elif material.stock <= 10:

            estado = "STOCK BAJO"

        else:

            estado = "NORMAL"

        ws.cell(fila, 1).value = material.codigo
        ws.cell(fila, 2).value = material.nombre
        ws.cell(fila, 3).value = material.unidad
        ws.cell(fila, 4).value = float(material.stock)
        ws.cell(fila, 5).value = estado

        fila += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="Inventario_General.xlsx"'

    wb.save(response)

    return response
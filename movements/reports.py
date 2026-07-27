from io import BytesIO

from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from .models import Delivery
from django.conf import settings
import os


def delivery_pdf(request, pk):

    """
    Genera el Vale de Entrega en formato PDF.
    """

    entrega = get_object_or_404(
        Delivery,
        pk=pk
    )

    buffer = BytesIO()

    documento = SimpleDocTemplate(

        buffer,

        pagesize=letter,

        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,

    )

    estilos = getSampleStyleSheet()

    estilo_titulo = estilos["Heading1"]
    estilo_titulo.alignment = TA_CENTER

    estilo_subtitulo = estilos["Heading2"]
    estilo_subtitulo.alignment = TA_CENTER

    estilo_normal = estilos["BodyText"]
    estilo_normal.alignment = TA_LEFT

    elementos = []

    # ----------------------------------------------------
    # ENCABEZADO
    # ----------------------------------------------------

    elementos.append(

        Paragraph(

            "<b>MATERIAL CONTROL CM</b>",

            estilo_titulo

        )

    )

    elementos.append(

        Paragraph(

            "VALE DE ENTREGA DE MATERIAL",

            estilo_subtitulo

        )

    )

    elementos.append(

        Spacer(
            1,
            0.50 * cm
        )

    )

    datos = [

        [
            "Entrega No.",
            str(entrega.id)
        ],

        [
            "Proyecto",
            str(entrega.proyecto)
        ],

        [
            "Responsable",
            entrega.responsable
        ],

        [
            "Fecha",
            entrega.fecha.strftime("%d/%m/%Y")
        ],

        [
            "Estado",
            entrega.get_estado_display()
        ],

    ]

    tabla_datos = Table(

        datos,

        colWidths=[
            5 * cm,
            11 * cm
        ]

    )

    tabla_datos.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (0, -1),
                colors.HexColor("#D9EAD3")
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),

            (
                "FONTNAME",
                (0, 0),
                (0, -1),
                "Helvetica-Bold"
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

        ])

    )

    elementos.append(tabla_datos)

    elementos.append(

        Spacer(
            1,
            0.70 * cm
        )

    )
        # ----------------------------------------------------
    # DETALLE DE MATERIALES
    # ----------------------------------------------------

    elementos.append(

        Paragraph(

            "<b>DETALLE DE MATERIALES</b>",

            estilos["Heading3"]

        )

    )

    elementos.append(

        Spacer(
            1,
            0.30 * cm
        )

    )

    encabezado = [

        [

            "Código",

            "Material",

            "Unidad",

            "Cantidad"

        ]

    ]

    detalle = []

    for item in entrega.detalles.select_related("material"):

        detalle.append(

            [

                item.material.codigo,

                item.material.nombre,

                item.material.unidad,

                str(item.cantidad),

            ]

        )

    if not detalle:

        detalle.append(

            [

                "-",

                "No existen materiales registrados.",

                "",

                ""

            ]

        )

    datos_tabla = encabezado + detalle

    tabla_materiales = Table(

        datos_tabla,

        colWidths=[

            3 * cm,

            8 * cm,

            2 * cm,

            3 * cm,

        ]

    )

    tabla_materiales.setStyle(

        TableStyle([

            (

                "BACKGROUND",

                (0, 0),

                (-1, 0),

                colors.HexColor("#0D6EFD")

            ),

            (

                "TEXTCOLOR",

                (0, 0),

                (-1, 0),

                colors.white

            ),

            (

                "FONTNAME",

                (0, 0),

                (-1, 0),

                "Helvetica-Bold"

            ),

            (

                "ALIGN",

                (2, 1),

                (3, -1),

                "CENTER"

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "MIDDLE"

            ),

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.5,

                colors.grey

            ),

            (

                "BOTTOMPADDING",

                (0, 0),

                (-1, 0),

                8

            ),

            (

                "BACKGROUND",

                (0, 1),

                (-1, -1),

                colors.beige

            ),

        ])

    )

    elementos.append(tabla_materiales)

    elementos.append(

        Spacer(

            1,

            0.70 * cm

        )

    )
        # ----------------------------------------------------
    # OBSERVACIONES
    # ----------------------------------------------------

    elementos.append(

        Paragraph(

            "<b>OBSERVACIONES</b>",

            estilos["Heading3"]

        )

    )

    elementos.append(

        Spacer(
            1,
            0.20 * cm
        )

    )

    observacion = entrega.observacion

    if not observacion:

        observacion = "Sin observaciones."

    tabla_obs = Table(

        [

            [observacion]

        ],

        colWidths=[16 * cm],

        rowHeights=[2.5 * cm]

    )

    tabla_obs.setStyle(

        TableStyle([

            (

                "GRID",

                (0, 0),

                (-1, -1),

                0.5,

                colors.black

            ),

            (

                "VALIGN",

                (0, 0),

                (-1, -1),

                "TOP"

            ),

            (

                "LEFTPADDING",

                (0, 0),

                (-1, -1),

                8

            ),

            (

                "TOPPADDING",

                (0, 0),

                (-1, -1),

                8

            ),

        ])

    )

    elementos.append(tabla_obs)

    elementos.append(

        Spacer(
            1,
            1 * cm
        )

    )

    # ----------------------------------------------------
    # FIRMAS
    # ----------------------------------------------------

    firmas = Table(

        [

            [

                "________________________",

                "________________________",

                "________________________",

            ],

            [

                "ENTREGÓ",

                "RECIBIÓ",

                "Vo. Bo.",

            ],

        ],

        colWidths=[5.3 * cm, 5.3 * cm, 5.3 * cm]

    )

    firmas.setStyle(

        TableStyle([

            (

                "ALIGN",

                (0, 0),

                (-1, -1),

                "CENTER"

            ),

            (

                "FONTNAME",

                (0, 1),

                (-1, 1),

                "Helvetica-Bold"

            ),

            (

                "TOPPADDING",

                (0, 1),

                (-1, 1),

                10

            ),

        ])

    )

    elementos.append(firmas)

    elementos.append(

        Spacer(
            1,
            1 * cm
        )

    )
        # ----------------------------------------------------
    # PIE DE PÁGINA
    # ----------------------------------------------------

    elementos.append(

        Paragraph(

            "<font size='8' color='grey'>"
            "Documento generado por Material Control CM"
            "</font>",

            estilo_normal

        )

    )

    # ----------------------------------------------------
    # GENERAR PDF
    # ----------------------------------------------------

    documento.build(

        elementos

    )

    pdf = buffer.getvalue()

    buffer.close()

    response = HttpResponse(

        content_type="application/pdf"

    )

    response["Content-Disposition"] = (

        f'inline; filename="Entrega_{entrega.id}.pdf"'

    )

    response.write(pdf)

    return response

def delivery_excel(request, pk):

    entrega = get_object_or_404(
        Delivery,
        pk=pk
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Vale de Entrega"

    # -------------------------------------------------
    # TÍTULO
    # -------------------------------------------------

    ws["A1"] = "MATERIAL CONTROL CM"
    ws["A2"] = "VALE DE ENTREGA DE MATERIAL"

    ws["A1"].font = Font(
        bold=True,
        size=16
    )

    ws["A2"].font = Font(
        bold=True,
        size=13
    )

    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    fila = 4
        # -------------------------------------------------
    # INFORMACIÓN DE LA ENTREGA
    # -------------------------------------------------

    ws[f"A{fila}"] = "Entrega No."
    ws[f"B{fila}"] = entrega.id
    fila += 1

    ws[f"A{fila}"] = "Proyecto"
    ws[f"B{fila}"] = str(entrega.proyecto)
    fila += 1

    ws[f"A{fila}"] = "Responsable"
    ws[f"B{fila}"] = entrega.responsable
    fila += 1

    ws[f"A{fila}"] = "Fecha"
    ws[f"B{fila}"] = entrega.fecha.strftime("%d/%m/%Y")
    fila += 1

    ws[f"A{fila}"] = "Estado"
    ws[f"B{fila}"] = entrega.get_estado_display()

    fila += 2

    # -------------------------------------------------
    # ENCABEZADOS DE LA TABLA
    # -------------------------------------------------

    encabezados = [

        "Código",

        "Material",

        "Unidad",

        "Cantidad",

    ]

    for columna, titulo in enumerate(encabezados, start=1):

        celda = ws.cell(
            row=fila,
            column=columna
        )

        celda.value = titulo

        celda.font = Font(
            bold=True,
            color="FFFFFF"
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="0D6EFD"
        )

        celda.alignment = Alignment(
            horizontal="center"
        )

        celda.border = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin"),

        )

    fila += 1

    # -------------------------------------------------
    # DETALLE DE MATERIALES
    # -------------------------------------------------

    for detalle in entrega.detalles.select_related("material"):

        ws.cell(
            row=fila,
            column=1
        ).value = detalle.material.codigo

        ws.cell(
            row=fila,
            column=2
        ).value = detalle.material.nombre

        ws.cell(
            row=fila,
            column=3
        ).value = detalle.material.unidad

        ws.cell(
            row=fila,
            column=4
        ).value = float(detalle.cantidad)

        for columna in range(1, 5):

            ws.cell(
                row=fila,
                column=columna
            ).border = Border(

                left=Side(style="thin"),

                right=Side(style="thin"),

                top=Side(style="thin"),

                bottom=Side(style="thin"),

            )

        fila += 1
        # -------------------------------------------------
    # AJUSTAR ANCHO DE COLUMNAS
    # -------------------------------------------------

    anchos = {
        "A": 18,
        "B": 45,
        "C": 15,
        "D": 15,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # -------------------------------------------------
    # OBSERVACIONES
    # -------------------------------------------------

    fila += 2

    ws[f"A{fila}"] = "Observaciones"
    ws[f"A{fila}"].font = Font(bold=True)

    fila += 1

    ws.merge_cells(
        start_row=fila,
        start_column=1,
        end_row=fila + 2,
        end_column=4
    )

    ws.cell(
        row=fila,
        column=1
    ).value = entrega.observacion or "Sin observaciones."

    ws.cell(
        row=fila,
        column=1
    ).alignment = Alignment(
        vertical="top",
        wrap_text=True
    )

    # -------------------------------------------------
    # GENERAR RESPUESTA
    # -------------------------------------------------

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = f'attachment; filename="Entrega_{entrega.id}.xlsx"'

    wb.save(response)

    return response

    return response
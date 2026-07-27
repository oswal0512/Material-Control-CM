from django.shortcuts import render
from django.http import HttpResponse
from materials.models import Material
from movements.models import InventoryMovement, Delivery
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph
)
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from movements.models import Delivery, DeliveryDetail, InventoryMovement
from openpyxl import Workbook
from django.http import HttpResponse
from accounts.decorators import gerencia_required

@gerencia_required
def report_home(request):

    print("=== REPORTING VIEWS ACTIVO ===")

    total_materiales = Material.objects.filter(
        activo=True
    ).count()

def report_home(request):

    total_materiales = Material.objects.filter(activo=True).count()
    total_movimientos = InventoryMovement.objects.count()
    total_entregas = Delivery.objects.count()

    materiales_bajo_stock = Material.objects.filter(
        stock__lte=10,
        activo=True
    )

    return render(
        request,
        "reporting/home.html",
        {
            "total_materiales": total_materiales,
            "total_movimientos": total_movimientos,
            "total_entregas": total_entregas,
            "materiales_bajo_stock": materiales_bajo_stock,
        }
    )


@gerencia_required
def inventory_pdf(request):

    materiales = Material.objects.filter(
        activo=True
    ).order_by(
        "nombre"
    )


    response = HttpResponse(
        content_type="application/pdf"
    )


    response["Content-Disposition"] = (
        'attachment; filename="inventario_actual.pdf"'
    )


    doc = SimpleDocTemplate(
        response,
        pagesize=letter
    )


    styles = getSampleStyleSheet()


    elementos = []


    titulo = Paragraph(
        "Material Control CM - Inventario Actual",
        styles["Title"]
    )


    elementos.append(titulo)


    data = [

        [
            "Código",
            "Material",
            "Unidad",
            "Stock"
        ]

    ]


    for material in materiales:

        data.append(
            [
                material.codigo,
                material.nombre,
                material.unidad,
                str(material.stock)
            ]
        )


    tabla = Table(data)


    tabla.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0,0),
                    (-1,0),
                    "grey"
                ),

                (
                    "GRID",
                    (0,0),
                    (-1,-1),
                    1,
                    "black"
                ),

            ]
        )
    )


    elementos.append(tabla)


    doc.build(elementos)


    return response
@gerencia_required
def deliveries_report(request):

    entregas = Delivery.objects.all().order_by(
        "-fecha"
    )


    proyecto = request.GET.get(
        "proyecto"
    )

    fecha_inicio = request.GET.get(
        "fecha_inicio"
    )

    fecha_fin = request.GET.get(
        "fecha_fin"
    )


    if proyecto:

        entregas = entregas.filter(
            proyecto_id=proyecto
        )


    if fecha_inicio:

        entregas = entregas.filter(
            fecha__gte=fecha_inicio
        )


    if fecha_fin:

        entregas = entregas.filter(
            fecha__lte=fecha_fin
        )


    from projects.models import Project


    proyectos = Project.objects.all()


    return render(
        request,
        "reporting/entregas.html",
        {
            "entregas": entregas,
            "proyectos": proyectos,
        }
    )
@gerencia_required
def inventory_excel(request):

    materiales = Material.objects.filter(
        activo=True
    ).order_by(
        "codigo"
    )


    wb = Workbook()

    ws = wb.active

    ws.title = "Inventario"


    ws.append([
        "Código",
        "Material",
        "Unidad",
        "Stock",
        "Estado"
    ])


    for material in materiales:

        estado = (
            "BAJO STOCK"
            if material.stock <= 10
            else "DISPONIBLE"
        )


        ws.append([
            material.codigo,
            material.nombre,
            material.unidad,
            float(material.stock),
            estado
        ])


    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="inventario.xlsx"'
    )


    wb.save(response)


    return response
@gerencia_required
def kardex_excel(request):

    movimientos = InventoryMovement.objects.all().order_by(
        "-fecha"
    )


    wb = Workbook()

    ws = wb.active

    ws.title = "Kardex"


    ws.append(
        [
            "Fecha",
            "Material",
            "Tipo",
            "Cantidad",
            "Saldo",
            "Referencia",
            "Responsable",
            "Observación",
        ]
    )


    for movimiento in movimientos:

        ws.append(
            [
                movimiento.fecha.strftime("%Y-%m-%d"),
                movimiento.material.nombre,
                movimiento.tipo,
                float(movimiento.cantidad),
                float(movimiento.saldo),
                movimiento.referencia,
                movimiento.responsable,
                movimiento.observacion,
            ]
        )


    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="kardex.xlsx"'
    )


    wb.save(response)


    return response
@gerencia_required
def deliveries_excel(request):

    detalles = DeliveryDetail.objects.select_related(
        "entrega",
        "material",
        "entrega__proyecto"
    ).order_by(
        "-entrega__fecha"
    )


    wb = Workbook()

    ws = wb.active

    ws.title = "Entregas"


    ws.append(
        [
            "Fecha",
            "Proyecto",
            "Responsable",
            "Material",
            "Cantidad",
            "Estado",
        ]
    )


    for detalle in detalles:

        ws.append(
            [
                detalle.entrega.fecha.strftime("%Y-%m-%d"),
                str(detalle.entrega.proyecto),
                detalle.entrega.responsable,
                detalle.material.nombre,
                float(detalle.cantidad),
                detalle.entrega.estado,
            ]
        )


    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="entregas_por_proyecto.xlsx"'
    )


    wb.save(response)


    return response
    
@gerencia_required
def management_report_pdf(request):

    materiales = Material.objects.filter(
        activo=True
    ).order_by(
        "nombre"
    )


    total_materiales = materiales.count()

    total_movimientos = InventoryMovement.objects.count()

    total_entregas = Delivery.objects.count()


    bajo_stock = Material.objects.filter(
        activo=True,
        stock__lte=10
    )


    response = HttpResponse(
        content_type="application/pdf"
    )


    response["Content-Disposition"] = (
        'attachment; filename="reporte_gerencial.pdf"'
    )


    doc = SimpleDocTemplate(
        response,
        pagesize=letter
    )


    styles = getSampleStyleSheet()


    elementos = []


    elementos.append(
        Paragraph(
            "Material Control CM",
            styles["Title"]
        )
    )


    elementos.append(
        Paragraph(
            "Reporte Gerencial de Inventario",
            styles["Heading2"]
        )
    )


    resumen = [

        ["Indicador", "Cantidad"],

        ["Materiales activos", total_materiales],

        ["Movimientos registrados", total_movimientos],

        ["Entregas realizadas", total_entregas],

        ["Materiales bajo stock", bajo_stock.count()],

    ]


    tabla_resumen = Table(
        resumen
    )


    tabla_resumen.setStyle(
        TableStyle(
            [
                (
                    "GRID",
                    (0,0),
                    (-1,-1),
                    1,
                    "black"
                ),

                (
                    "BACKGROUND",
                    (0,0),
                    (-1,0),
                    "grey"
                ),
            ]
        )
    )


    elementos.append(
        tabla_resumen
    )


    elementos.append(
        Paragraph(
            "Materiales con bajo stock",
            styles["Heading2"]
        )
    )


    datos_stock = [

        [
            "Código",
            "Material",
            "Stock"
        ]

    ]


    for material in bajo_stock:

        datos_stock.append(
            [
                material.codigo,
                material.nombre,
                str(material.stock)
            ]
        )


    tabla_stock = Table(
        datos_stock
    )


    tabla_stock.setStyle(
        TableStyle(
            [
                (
                    "GRID",
                    (0,0),
                    (-1,-1),
                    1,
                    "black"
                ),

                (
                    "BACKGROUND",
                    (0,0),
                    (-1,0),
                    "grey"
                ),
            ]
        )
    )


    elementos.append(
        tabla_stock
    )


    doc.build(
        elementos
    )


    return response

    elementos.append(tabla)


    doc.build(elementos)


    return response